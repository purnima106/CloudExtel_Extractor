from fastapi import FastAPI, File, UploadFile, Form, APIRouter, HTTPException, Request, BackgroundTasks, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from typing import List, Optional, Dict
import io
import os
import tempfile
import json
import traceback
import uuid
import zipfile
import threading
import pandas as pd
from datetime import datetime
from supabase import create_client, Client
#from extract_trench_data import process_demand_note, append_row_to_excel, sanitize_filename
import re
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import numpy as np
import math
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Import and include the actual_cost_extraction router
from parsers.actual_cost_extraction import router as actual_cost_extraction_router
# from parsers.dn_master_upload import router as dn_master_upload_router
from dotenv import load_dotenv
from parsers.mcgm_application_parser import mcgm_application_parser
from parsers.universal_application_parser import universal_application_parser
import fitz
from parsers.permit_parser import extract_permit_fields, upsert_permit_fields_to_dn_master
from parsers.clientparserv2 import unified_parser, generate_non_refundable_output, generate_sd_output

# Add these imports at the top of the file (after other imports):
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer, PageBreak, Frame, PageTemplate
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import landscape

load_dotenv()

import os
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")

if SUPABASE_KEY:
    print(f"[LOG] Loaded SUPABASE_KEY: {SUPABASE_KEY[:8]}...{SUPABASE_KEY[-4:]}")
else:
    print("[ERROR] SUPABASE_KEY not found in environment!")

app = FastAPI()

# Allow CORS for local frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],  # Allow only frontend origin
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition", "content-disposition"],  # <-- Expose for frontend JS
)

# Global in-memory cache for parsed preview data
preview_cache = {}
preview_cache_lock = threading.Lock()
CACHE_EXPIRY_SECONDS = 3600  # 1 hour (optional, not enforced in this snippet)

# Include the actual_cost_extraction router
app.include_router(actual_cost_extraction_router)
# app.include_router(dn_master_upload_router)

DN_MASTER_COLUMNS = [
    "sr_no", "route_type", "ip1_co_built", "dn_recipient", "project_name", "site_id", "uid",
    "build_type", "category_type", "po_number", "po_length", "route_id_lmc_id",
    "parent_route", "route_lmc_id", "route_lmc_section_id", "route_lmc_subsection_id", "application_number",
    "application_length_mtr", "application_date", "from_location", "to_location", "authority", "ward",
    "survey_done_mtr", "dn_number", "dn_length_mtr", "dn_received_date", "type", "trench_type", "ot_hdd", "pit",
    "surface", "surface_wise_ri_amount", "dn_ri_amount", "ground_rent", "administrative_charge",
    "supervision_charges", "chamber_fee", "gst", "deposit", "total_dn_amount", "ri_budget_amount_per_meter",
    "projected_budget_ri_amount_dn", "actual_total_non_refundable", "non_refundable_amount_per_mtr",
    "proj_non_refundable_savings_per_mtr", "deposit_repeat", "total_dn_amount_repeat", "row_network_id",
    "route_network_id", "new_revised_dn_number", "new_revised_dn_against", "internal_approval_start",
    "internal_approval_end", "ticket_raised_date", "dn_payment_date", "tat_days", "civil_completion_date",
    "road_name"
]

# Add this mapping at the top (after DN_MASTER_COLUMNS or near SUPABASE_URL)
FIELD_MAP = {
    "sr_no": "sr_no",
    "route_type": "route_type",
    "route_routeLM_metroLM_LMCStandalone": "route_routeLM_metroLM_LMCStandalone",
    "ip1_co_built": "ip1_co_built",
    "dn_recipient": "dn_recipient",
    "project_name": "project_name",
    "route_id / site_id": "route_id_site_id",
    "SiteID": "route_id_site_id",
    "route_id_site_id": "route_id_site_id",
    "uid": "uid",
    "build_type": "build_type",
    "category_type": "category_type",
    # Removed survey_id - no longer needed with one budget per route
    "po_number": "po_number",
    "PO No": "po_number",
    "po_length": "po_length",
    "PO Length (Mtr)": "po_length",
    "parent_route": "parent_route",
    "Parent Route Name / HH": "parent_route",
    "ce_route_lmc_id": "ce_route_lmc_id",
    "route_lmc_section_id": "route_lmc_section_id",
    "route_lmc_subsection_id": "route_lmc_subsection_id",
    "application_number": "application_number",
    "Application Number": "application_number",
    "application_length_mtr": "application_length_mtr",
    "Application Length (Mtr)": "application_length_mtr",
    "application_date": "application_date",
    "Application Date": "application_date",
    "from_location": "from_location",
    "From": "from_location",
    "to_location": "to_location",
    "To": "to_location",
    "authority": "authority",
    "Authority": "authority",
    "ward": "ward",
    "Ward": "ward",
    "dn_number": "dn_number",
    "Demand Note Reference number": "dn_number",
    "dn_length_mtr": "dn_length_mtr",
    "Section Length": "dn_length_mtr",
    "dn_received_date": "dn_received_date",
    "DN Received Date": "dn_received_date",
    "trench_type": "trench_type",
    "ot_length": "ot_length",
    "surface": "surface",
    "Road Types": "surface",
    "surface_wise_ri_amount": "surface_wise_ri_amount",
    "ri_rate_go_rs": "surface_wise_ri_amount",
    "Rate in Rs": "surface_wise_ri_amount",
    "dn_ri_amount": "dn_ri_amount",
    "RI Amount": "dn_ri_amount",
    "multiplying_factor": "multiplying_factor",
    "Multiplication Factor": "multiplying_factor",
    "ground_rent": "ground_rent",
    "Ground Rent": "ground_rent",
    "administrative_charge": "administrative_charge",
    "Administrative Charge": "administrative_charge",
    "supervision_charges": "supervision_charges",
    "Supervision Charges": "supervision_charges",
    "chamber_fee": "chamber_fee",
    "Chamber Fee": "chamber_fee",
    "gst": "gst",
    "GST Amount": "gst",
    "ri_budget_amount_per_meter": "ri_budget_amount_per_meter",
    "projected_budget_ri_amount_dn": "projected_budget_ri_amount_dn",
    "actual_total_non_refundable": "actual_total_non_refundable",
    "non_refundable_amount_per_mtr": "non_refundable_amount_per_mtr",
    "proj_non_refundable_savings_per_mtr": "proj_non_refundable_savings_per_mtr",
    "proj_savings_per_dn": "proj_savings_per_dn",
    "deposit": "deposit",
    "SD Amount": "deposit",
    "total_dn_amount": "total_dn_amount",
    "Total DN Amount": "total_dn_amount",
    "new_revised_dn_number": "new_revised_dn_number",
    "new_revised_dn_against": "new_revised_dn_against",
    "internal_approval_start": "internal_approval_start",
    "internal_approval_end": "internal_approval_end",
    "ticket_raised_date": "ticket_raised_date",
    "dn_payment_date": "dn_payment_date",
    "tat_days": "tat_days",
    "civil_completion_date": "civil_completion_date",
    "hdd_length": "hdd_length",
    "no_of_pits": "no_of_pits",
    "pit_ri_rate": "pit_ri_rate",
    "surface_wise_length": "surface_wise_length",
    "surface_wise_ri_amount": "surface_wise_ri_amount",
    "surface_wise_multiplication_factor": "surface_wise_multiplication_factor",
    "Road Name": "road_name",
    "road_name": "road_name",
    # Add more mappings as needed
}

# Add this near the top, after FIELD_MAP or DN_MASTER_COLUMNS
VALIDATE_PARSER_FIELDS = [
    'sr_no', 'route_type', 'ip1_co_built', 'dn_recipient', 'project_name', 'route_id_site_id', 'uid',
    'build_type', 'category_type', 'po_number', 'po_length', 'parent_route', 'ce_route_lmc_id',
    'route_lmc_section_id', 'route_lmc_subsection_id', 'application_number', 'application_length_mtr', 'application_date',
    'from_location', 'to_location', 'authority', 'ward', 'dn_number', 'dn_length_mtr', 'dn_received_date', 'trench_type',
    'ot_length', 'surface', 'surface_wise_ri_amount', 'dn_ri_amount', 'ground_rent', 'administrative_charge',
    'supervision_charges', 'chamber_fee', 'gst', 'ri_budget_amount_per_meter', 'projected_budget_ri_amount_dn',
    'actual_total_non_refundable', 'non_refundable_amount_per_mtr', 'proj_non_refundable_savings_per_mtr', 'deposit',
    'total_dn_amount', 'new_revised_dn_number', 'new_revised_dn_against', 'internal_approval_start', 'internal_approval_end',
    'ticket_raised_date', 'dn_payment_date', 'tat_days', 'civil_completion_date', 'hdd_length', 'no_of_pits', 'pit_ri_rate',
    'proj_savings_per_dn',
    'surface_wise_length',
    'surface_wise_ri_amount',
    'surface_wise_multiplication_factor',
    'road_name',
]

# Normalization function for field names
def normalize_field_name(name):
    # Lowercase, remove spaces, replace underscores and slashes with nothing
    return re.sub(r'[\s_\/]+', '', name).lower()

# Canonical allowed fields (normalized)
NORMALIZED_ALLOWED_FIELDS = {normalize_field_name(f): f for f in VALIDATE_PARSER_FIELDS}

# Use the new DB column name for route_id_site_id
ROUTE_ID_SITE_ID_CANONICAL = 'route_id_site_id'

# List of date fields in the DB
DATE_FIELDS = {
    'application_date',
    'dn_received_date',
    'internal_approval_start',
    'internal_approval_end',
    'ticket_raised_date',
    'dn_payment_date',
    'civil_completion_date'
}

def normalize_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, str):
        # Try DD/MM/YYYY
        try:
            return datetime.strptime(val, '%d/%m/%Y').strftime('%Y-%m-%d')
        except Exception:
            pass
        # Try YYYY-MM-DD (already correct)
        try:
            return datetime.strptime(val, '%Y-%m-%d').strftime('%Y-%m-%d')
        except Exception:
            pass
    return None  # Return None if not a recognized date string

def fetch_ri_cost_per_meter_from_supabase(site_id):
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    # Always use the correct table and columns for budget lookup
    response = supabase.table("budget_master").select("ri_cost_per_meter").eq("route_id_site_id", site_id).execute()
    if response.data and len(response.data) > 0:
        return response.data[0].get("ri_cost_per_meter", "")
    return ""

@app.post("/process")
async def process_pdf(
    authority: str = Form(...),
    manual_fields: Optional[str] = Form(None),  # JSON string of manual fields (Non-Refundable)
    sd_manual_fields: Optional[str] = Form(None),  # JSON string of manual fields (SD Output)
    file: UploadFile = File(...)
):
    # Save uploaded file to a temp location (cross-platform, ensure unique name)
    temp_dir = tempfile.gettempdir()
    unique_filename = f"{uuid.uuid4()}_{file.filename}"
    temp_path = os.path.join(temp_dir, unique_filename)
    file_bytes = await file.read()
    with open(temp_path, "wb") as f:
        f.write(file_bytes)
    print(f"[DEBUG] Saved file: {temp_path}, size: {os.path.getsize(temp_path)} bytes, first 8 bytes: {file_bytes[:8]}")

    # Parse manual fields if provided
    manual_fields_dict = json.loads(manual_fields) if manual_fields else {}
    sd_manual_fields_dict = json.loads(sd_manual_fields) if sd_manual_fields else {}
    # Call extraction logic, get both file paths
    try:
        non_ref_xlsx_path, sd_xlsx_path = process_demand_note(temp_path, authority, manual_fields_dict, sd_manual_fields_dict, return_paths=True)
        # Create a zip with both files
        zip_path = os.path.join(temp_dir, f"{uuid.uuid4()}_outputs.zip")
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            zipf.write(non_ref_xlsx_path, arcname=os.path.basename(non_ref_xlsx_path))
            zipf.write(sd_xlsx_path, arcname=os.path.basename(sd_xlsx_path))
        os.remove(temp_path)
        return FileResponse(zip_path, filename="outputs.zip", media_type="application/zip")
    except Exception as e:
        traceback.print_exc()
        os.remove(temp_path)
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/process/non_refundable")
async def process_non_refundable(
    background_tasks: BackgroundTasks,
    authority: str = Form(...),
    manual_fields: Optional[str] = Form(None),
    file: UploadFile = File(None),
    preview_id: Optional[str] = Form(None)
):
    import tempfile, os, uuid, json, traceback
    temp_dir = tempfile.gettempdir()
    manual_fields_dict = json.loads(manual_fields) if manual_fields else {}
    try:
        # Strictly require preview cache
        if not preview_id or preview_id not in preview_cache:
            return JSONResponse(status_code=400, content={"error": "Preview data not found. Please preview the file first."})
        cached = preview_cache[preview_id]
        row = cached['row']
        headers = cached['headers']
        demand_note_number = cached.get('demand_note_number', 'Output')
        # Update cached row with latest manual fields before writing Excel
        if manual_fields_dict:
            for field, value in manual_fields_dict.items():
                if field in headers:
                    idx = headers.index(field)
                    row[idx] = value
        from extract_trench_data import append_row_to_excel
        temp_excel_path = os.path.join(temp_dir, f"{uuid.uuid4()}_Non_Refundable_Output.xlsx")
        # Determine blue_headers for authority
        if authority.upper() == "MCGM":
            blue_headers = [
                "LM/BB/FTTH", "GO RATE", "Total Route (MTR)", "Not part of capping (License Fee/Rental Payment /Way Leave charges etc.)",
                "REASON FOR DELAY (>2 DAYS)", "PO No.", "Route Name(As per CWIP)", "Section Name for ROW(As per CWIP)"
            ]
        elif authority.upper() == "MBMC":
            blue_headers = [
                "LM/BB/FTTH", "GO RATE", "Total Route (MTR)", "Not part of capping (License Fee/Rental Payment /Way Leave charges etc.)",
                "REASON FOR DELAY (>2 DAYS)", "PO No.", "Route Name(As per CWIP)", "Section Name for ROW(As per CWIP)"
            ]
        elif authority.upper() == "NMMC":
            blue_headers = []
        else:
            blue_headers = []
        append_row_to_excel(temp_excel_path, row, headers, manual_fields=manual_fields_dict, blue_headers=blue_headers)
        download_filename = f"{sanitize_filename(demand_note_number)}_Non Refundable Output.xlsx"
        from fastapi.responses import FileResponse
        if background_tasks is not None:
            background_tasks.add_task(os.remove, temp_excel_path)
        return FileResponse(
            temp_excel_path,
            filename=download_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/process/sd")
async def process_sd(
    background_tasks: BackgroundTasks,
    authority: str = Form(...),
    sd_manual_fields: Optional[str] = Form(None),
    file: UploadFile = File(None),
    preview_id: Optional[str] = Form(None)
):
    import tempfile, os, uuid, json, traceback
    temp_dir = tempfile.gettempdir()
    sd_manual_fields_dict = json.loads(sd_manual_fields) if sd_manual_fields else {}
    try:
        # Strictly require preview cache
        if not preview_id or preview_id not in preview_cache:
            return JSONResponse(status_code=400, content={"error": "Preview data not found. Please preview the file first."})
        cached = preview_cache[preview_id]
        row = cached['row']
        headers = cached['headers']
        demand_note_number = cached.get('demand_note_number', 'Output')
        # Update cached row with latest manual fields before writing Excel
        if sd_manual_fields_dict:
            for field, value in sd_manual_fields_dict.items():
                if field in headers:
                    idx = headers.index(field)
                    row[idx] = value
        from extract_trench_data import append_row_to_excel
        temp_excel_path = os.path.join(temp_dir, f"{uuid.uuid4()}_SD_Output.xlsx")
        # Determine blue_headers for authority
        if authority.upper() == "MCGM":
            blue_headers = [
                "Execution Partner GBPA PO No.", "Partner PO circle", "Unique route id", "NFA no."
            ]
        elif authority.upper() == "MBMC":
            blue_headers = [
                "Execution Partner GBPA PO No.", "Partner PO circle", "Unique route id", "NFA no."
            ]
        elif authority.upper() == "NMMC":
            blue_headers = []
        else:
            blue_headers = []
        append_row_to_excel(temp_excel_path, row, headers, manual_fields=sd_manual_fields_dict, blue_headers=blue_headers)
        download_filename = f"{sanitize_filename(demand_note_number)}_SD Output.xlsx"
        from fastapi.responses import FileResponse
        if background_tasks is not None:
            background_tasks.add_task(os.remove, temp_excel_path)
        return FileResponse(
            temp_excel_path,
            filename=download_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/preview/non_refundable")
async def preview_non_refundable(
    authority: str = Form(...),
    manualFields: Optional[str] = Form(None),
    file: UploadFile = File(...)
):
    import tempfile, os, json, traceback, uuid
    global preview_cache, preview_cache_lock
    manual_fields_dict = json.loads(manualFields) if manualFields else {}
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(await file.read())
            tmp_path = tmp.name
        try:
            row = None
            headers = None
            demand_note_number = None
            if authority.upper() == "MCGM":
                from parsers.mcgm import non_refundable_request_parser
                from constants.comprehensive_field_mapping import get_comprehensive_field_mapping
                row = non_refundable_request_parser(tmp_path, manual_values=manual_fields_dict)
                mapping = get_comprehensive_field_mapping()
                headers = mapping["all_validation_fields"]

                if row is None:
                    print("[MCGM ERROR] Parser returned None for MCGM.")
                    return JSONResponse(status_code=400, content={"error": "Parser returned no data for MCGM. Check if the uploaded file is correct and supported."})
                if not isinstance(row, dict):
                    print(f"[MCGM ERROR] Parser did not return a dict. Type: {type(row)} Value: {row}")
                    return JSONResponse(status_code=400, content={"error": "Parser did not return a dict for MCGM."})
                
                # Use the row data directly (it already contains static values)
                preview_data = row
                
                # For MCGM, the row is already in non-refundable format with display names
                # So we need to use the non-refundable field names for cache
                from constants.comprehensive_field_mapping import ALL_NON_REFUNDABLE_FIELDS
                row_for_cache = [preview_data.get(h, "") for h in ALL_NON_REFUNDABLE_FIELDS]

                demand_note_number = preview_data.get("Demand Note Reference number", "Output")
                preview_id = str(uuid.uuid4())
                with preview_cache_lock:
                    preview_cache[preview_id] = {
                        'row': row_for_cache,
                        'headers': ALL_NON_REFUNDABLE_FIELDS,
                        'demand_note_number': demand_note_number
                    }

                return {"rows": [preview_data], "preview_id": preview_id}
                
            elif authority.upper() == "MBMC":
                from parsers.mbmc import non_refundable_request_parser
                from constants.comprehensive_field_mapping import get_comprehensive_field_mapping
                row = non_refundable_request_parser(tmp_path, manual_values=manual_fields_dict)
                mapping = get_comprehensive_field_mapping()
                headers = mapping["all_validation_fields"]

                if row is None:
                    print("[MBMC ERROR] Parser returned None for MBMC.")
                    return JSONResponse(status_code=400, content={"error": "Parser returned no data for MBMC. Check if the uploaded file is correct and supported."})
                if not isinstance(row, dict):
                    print(f"[MBMC ERROR] Parser did not return a dict. Type: {type(row)} Value: {row}")
                    return JSONResponse(status_code=400, content={"error": "Parser did not return a dict for MBMC."})
                
                # Use the row data directly (it already contains static values)
                preview_data = row
                
                row_for_cache = [preview_data.get(h, "") for h in headers]

                demand_note_number = preview_data.get("Demand Note Reference number", "Output")
                preview_id = str(uuid.uuid4())
                with preview_cache_lock:
                    preview_cache[preview_id] = {
                        'row': row_for_cache,
                        'headers': headers,
                        'demand_note_number': demand_note_number
                    }

                return {"rows": [preview_data], "preview_id": preview_id}
                
            elif authority.upper() == "KDMC":
                from parsers.kdmc import non_refundable_request_parser, PREVIEW_NON_REFUNDABLE_COLUMNS
                row = non_refundable_request_parser(tmp_path, manual_values=manual_fields_dict)
                headers = PREVIEW_NON_REFUNDABLE_COLUMNS

                if row is None:
                    print("[KDMC ERROR] Parser returned None for KDMC.")
                    return JSONResponse(status_code=400, content={"error": "Parser returned no data for KDMC. Check if the uploaded file is correct and supported."})
                if not isinstance(row, dict):
                    print(f"[KDMC ERROR] Parser did not return a dict. Type: {type(row)} Value: {row}")
                    return JSONResponse(status_code=400, content={"error": "Parser did not return a dict for KDMC."})
                if not headers or not isinstance(headers, list) or len(headers) == 0:
                    print("[KDMC ERROR] Headers are missing or empty for KDMC.")
                    return JSONResponse(status_code=400, content={"error": "No headers found for KDMC."})
                
                # Use the row data directly (it already contains static values)
                preview_data = row
                
                row_for_cache = [preview_data.get(h, "") for h in headers]

                demand_note_number = preview_data.get("Demand Note Reference number", "Output")
                preview_id = str(uuid.uuid4())
                with preview_cache_lock:
                    preview_cache[preview_id] = {
                        'row': row_for_cache,
                        'headers': headers,
                        'demand_note_number': demand_note_number
                    }

                return {"rows": [preview_data], "preview_id": preview_id}
            else:
                return JSONResponse(status_code=400, content={"error": "Preview not implemented for this authority"})
        finally:
            os.remove(tmp_path)
    except Exception as e:
        print("[KDMC ERROR] Exception in preview_non_refundable:", e)
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/preview/sd")
async def preview_sd(
    authority: str = Form(...),
    manualFields: Optional[str] = Form(None),
    file: UploadFile = File(...)
):
    import tempfile, os, json, traceback, uuid
    manual_fields_dict = json.loads(manualFields) if manualFields else {}
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(await file.read())
            tmp_path = tmp.name
        try:
            alt_headers = None
            row_alt = None
            demand_note_number = None
            if authority.upper() == "MCGM":
                from parsers.mcgm import sd_parser
                alt_headers, row_alt = sd_parser(tmp_path, manual_values=manual_fields_dict)
            elif authority.upper() == "MBMC":
                from parsers.mbmc import sd_parser
                alt_headers, row_alt = sd_parser(tmp_path, manual_values=manual_fields_dict)
            elif authority.upper() == "KDMC":
                from parsers.kdmc import sd_parser, PREVIEW_SD_COLUMNS
                alt_headers, row_alt = sd_parser(tmp_path, manual_values=manual_fields_dict)
                
                # Create preview data with all standard SD columns
                preview_data = {}
                for i, col in enumerate(alt_headers):
                    preview_data[col] = row_alt[i] if i < len(row_alt) else ""
                
                # Ensure all expected SD columns are present (double-check)
                for col in PREVIEW_SD_COLUMNS:
                    if col not in preview_data:
                        preview_data[col] = ""
                
                demand_note_number = preview_data.get("DN No", "Output")
                preview_id = str(uuid.uuid4())
                with preview_cache_lock:
                    preview_cache[preview_id] = {
                        'row': [preview_data.get(h, "") for h in alt_headers],
                        'headers': alt_headers,
                        'demand_note_number': demand_note_number
                    }

                return {"rows": [preview_data], "preview_id": preview_id}
            else:
                return JSONResponse(status_code=400, content={"error": "Preview not implemented for this authority"})
            preview_data = {h: row_alt[i] for i, h in enumerate(alt_headers)}
            demand_note_number = preview_data.get("DN No", "Output")
            preview_id = str(uuid.uuid4())
            with preview_cache_lock:
                preview_cache[preview_id] = {
                    'row': row_alt,
                    'headers': alt_headers,
                    'demand_note_number': demand_note_number
                }
            
            return {"rows": [preview_data], "preview_id": preview_id}
        finally:
            os.remove(tmp_path)
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.options("/process/non_refundable")
async def options_non_refundable():
    from fastapi.responses import Response
    print("[DEBUG] [CORS] Preflight OPTIONS /process/non_refundable")
    return Response(status_code=204)

@app.options("/process/sd")
async def options_sd():
    from fastapi.responses import Response
    print("[DEBUG] [CORS] Preflight OPTIONS /process/sd")
    return Response(status_code=204)

@app.get("/debug/headers")
def debug_headers():
    from fastapi import Request
    from fastapi.responses import JSONResponse
    def _headers_to_dict(headers):
        return {k: v for k, v in headers.items()}
    # This endpoint is for manual curl/browser testing
    return JSONResponse({
        "request_headers": dict(),  # Not available in GET, but placeholder
        "note": "Check browser network tab for response headers."
    })

@app.get("/")
def root():
    return {"status": "FastAPI backend running"}

@app.post("/api/parse-application")
async def parse_application_file(
    authority: str = Form(None),
    dn_application_file: UploadFile = File(...)
):
    if not dn_application_file:
        return JSONResponse(status_code=422, content={"error": "Missing dn_application_file"})
    import tempfile, os
    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, f"dn_app_{dn_application_file.filename}")
    file_bytes = await dn_application_file.read()
    with open(temp_path, "wb") as f:
        f.write(file_bytes)
    try:
        if authority.upper() == "MCGM":
            result = mcgm_application_parser(temp_path)
        else:
            from parsers.universal_application_parser import universal_application_parser
            parsed = universal_application_parser(temp_path)
            os.remove(temp_path)
            return parsed if isinstance(parsed, dict) else {"result": parsed}
        os.remove(temp_path)
        return result if isinstance(result, dict) else {"result": result}
    except Exception as e:
        os.remove(temp_path)
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/parse-po")
async def parse_po_db(site_id: str = Form(...), po_number_type: str = Form(None)):
    from supabase import create_client
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    print(f"🔍 PO Parse API called with site_id: {site_id}, po_number_type: {po_number_type}")
    
    # Query po_master for the matching row
    response = supabase.table("po_master").select("*").eq("route_id_site_id", site_id).execute()
    if not response.data or len(response.data) == 0:
        return {"error": "No matching row found in po_master."}
    row = response.data[0]
    
    print(f"📊 Database row found: {row}")
    
    def clean_value(val):
        if val is None or str(val).strip() in ('', '-', 'nan', 'None'):
            return ""
        return str(val).strip()
    
    # PO No logic - Use frontend PO Number Type selection instead of route_type
    if po_number_type and po_number_type.lower() == "ip1":
        po_no = clean_value(row.get('po_no_ip1', ""))
        po_length = clean_value(row.get('po_length_ip1', ""))
        print(f"🎯 Using IP1 fields: po_no_ip1='{po_no}', po_length_ip1='{po_length}'")
    elif po_number_type and po_number_type.lower() == "co-built":
        po_no = clean_value(row.get('po_no_cobuild', ""))
        po_length = clean_value(row.get('po_length_cobuild', ""))
        print(f"🎯 Using Co-Build fields: po_no_cobuild='{po_no}', po_length_cobuild='{po_length}'")
    else:
        # Fallback to route_type logic if no PO Number Type provided
        route_type_val = clean_value(row.get('route_type', ""))
        route_type_val_norm = route_type_val.replace(" ", "").lower()
        print(f"⚠️ No PO Number Type provided, falling back to route_type: '{route_type_val}' (normalized: '{route_type_val_norm}')")
        if route_type_val_norm in ["metrolm", "lmc(standalone)", "routelm"]:
            po_no = clean_value(row.get('po_no_cobuild', ""))
            po_length = clean_value(row.get('po_length_cobuild', ""))
            print(f"🔄 Fallback: Using Co-Build fields based on route_type")
        elif route_type_val_norm == "route":
            po_no = clean_value(row.get('po_no_ip1', ""))
            po_length = clean_value(row.get('po_length_ip1', ""))
            print(f"🔄 Fallback: Using IP1 fields based on route_type")
        else:
            po_no = ""
            po_length = ""
            print(f"❌ Fallback: Unknown route_type, setting empty PO fields")
    
    # Category: always from 'route_type'
    category_val = clean_value(row.get('route_type', ""))
    # UID: always from 'uid'
    uid_val = clean_value(row.get('uid', ""))
    # Parent Route Name / HH: always from 'parent_route'
    parent_route_val = clean_value(row.get('parent_route', ""))
    
    result = {
        'PO No': po_no,
        'PO Length (Mtr)': po_length,
        'Category': category_val,
        'SiteID': clean_value(row.get('route_id_site_id', site_id)),
        'UID': uid_val,
        'Parent Route Name / HH': parent_route_val
    }
    
    print(f"📤 Returning result: {result}")
    return result

@app.post("/api/parse-dn")
async def parse_dn_file(
    authority: str = Form(...),
    dn_file: UploadFile = File(None),
    file: UploadFile = File(None),
    site_id: str = Form(None)  # <-- Accept site_id from frontend
):
    def clean_value(val):
        if val is None or str(val).strip() in ('', '-', 'nan', 'None'):
            return ""
        return str(val).strip()
    actual_file = dn_file or file

    import tempfile, os
    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, f"dn_{getattr(actual_file, 'filename', 'nofile')}")
    if actual_file is None:

        return JSONResponse(status_code=400, content={"error": "No file provided."})
    file_bytes = await actual_file.read()
    with open(temp_path, "wb") as f:
        f.write(file_bytes)
    try:
        # --- Parse the DN file for any authority ---
        if authority.upper() == "MBMC":
            from parsers.mbmc import non_refundable_request_parser, HEADERS
            row = non_refundable_request_parser(temp_path)
            headers = HEADERS
            if isinstance(row, dict):
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                dn_row = row
            else:
                dn_row = {h: row[i] for i, h in enumerate(headers)}
                if os.path.exists(temp_path):
                    os.remove(temp_path)
        elif authority.upper() == "MCGM":
            from parsers.mcgm import extract_all_fields_for_testing
            dn_row = extract_all_fields_for_testing(temp_path)
            if os.path.exists(temp_path):
                os.remove(temp_path)
        elif authority.upper() == "NMMC":
            from parsers.nmmc import extract_nmmc_all_fields
            # Extract raw fields and use comprehensive field mapping
            extracted_fields = extract_nmmc_all_fields(temp_path)
            if os.path.exists(temp_path):
                os.remove(temp_path)
            
            # Use comprehensive field mapping to convert to standard field names
            from constants.comprehensive_field_mapping import map_parser_to_standard
            dn_row = map_parser_to_standard(extracted_fields, "nmmc")
        elif authority.upper() == "KDMC":
            from parsers.kdmc import extract_kdmc_all_fields
            # Extract raw fields and use comprehensive field mapping
            extracted_fields = extract_kdmc_all_fields(temp_path)
            if os.path.exists(temp_path):
                os.remove(temp_path)
            
            # Use comprehensive field mapping to convert to standard field names
            from constants.comprehensive_field_mapping import map_parser_to_standard
            dn_row = map_parser_to_standard(extracted_fields, "kdmc")
        else:
            if os.path.exists(temp_path):
                os.remove(temp_path)
    
            return JSONResponse(status_code=400, content={"error": f"Unsupported authority: {authority}"})

        # --- PATCH: Enrich with PO, budget, and calculated fields for ALL authorities ---
        used_site_id = site_id if site_id else dn_row.get('route_id_site_id', '')
        dn_row['route_id_site_id'] = used_site_id
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        # Fetch PO row
        po_row = None
        response = supabase.table("po_master").select("*").eq("route_id_site_id", used_site_id).execute()
        if response.data and len(response.data) > 0:
            po_row = response.data[0]

        # Fetch budget fields for build_type, category_type, and ri_cost_per_meter
        budget_row = None
        try:
            budget_resp = supabase.table("budget_master").select("build_type", "category_type", "ri_cost_per_meter").eq("route_id_site_id", used_site_id).execute()
            if budget_resp.data and len(budget_resp.data) > 0:
                budget_row = budget_resp.data[0]
        except Exception as e:
            pass
        
        # Get route_type from PO master instead of budget master
        route_type_po = clean_value(po_row.get('route_type', '')) if po_row else ''
        # Determine if this is a route using po_master.route_type
        route_type_po_norm = route_type_po.replace(" ","").replace("\n","").lower()
        is_route = route_type_po_norm == "route"

        # Get user input for PO Number Type (IP1 or Co-Built)
        po_number_type = ''
        try:
            # Try to get from request form (if available)
            import starlette.requests
            if isinstance(locals().get('request'), starlette.requests.Request):
                form = await request.form()
                po_number_type = form.get('po_number_type') or form.get('ip1_co_built_input') or ''
        except Exception:
            pass
        # Fallback: try to get from request_data if present
        if not po_number_type and 'request_data' in locals():
            po_number_type = request_data.get('po_number_type', '') or request_data.get('ip1_co_built_input', '')

        # PO logic for all authorities
        if not is_route:
            ip1_co_built_val = "Co-Built"
            po_no = clean_value(po_row.get('po_no_cobuild', "")) if po_row else ''
            po_length = clean_value(po_row.get('po_length_cobuild', "")) if po_row else ''
            dn_recipient_val = "CE"
        else:
            ip1_co_built_val = po_number_type or 'IP1'
            if ip1_co_built_val.lower() == "ip1":
                po_no = clean_value(po_row.get('po_no_ip1', "")) if po_row else ''
                po_length = clean_value(po_row.get('po_length_ip1', "")) if po_row else ''
                dn_recipient_val = "Airtel"
            else:
                po_no = clean_value(po_row.get('po_no_cobuild', "")) if po_row else ''
                po_length = clean_value(po_row.get('po_length_cobuild', "")) if po_row else ''
                dn_recipient_val = "CE"

        # project_name always hardcoded
        project_name_val = "Mumbai Fiber Refresh Project"
        # route_id_site_id always from user input
        route_id_site_id_val = used_site_id
        # build_type/category_type from budget_master
        build_type_val = budget_row["build_type"] if budget_row and "build_type" in budget_row else ''
        category_type_val = budget_row["category_type"] if budget_row and "category_type" in budget_row else ''
        # po_number/po_length logic based on ip1_co_built
        if ip1_co_built_val.lower() == "co-built":
            po_no = clean_value(po_row.get('po_no_cobuild', "")) if po_row else ''
            po_length = clean_value(po_row.get('po_length_cobuild', "")) if po_row else ''
        else:
            po_no = clean_value(po_row.get('po_no_ip1', "")) if po_row else ''
            po_length = clean_value(po_row.get('po_length_ip1', "")) if po_row else ''

        # --- ENSURE route_type_val is always defined ---
        route_type_val = route_type_po  # Always set from po_master
        # Removed survey_id_val - no longer needed with one budget per route
        # --- ENSURE po_no is always a string and never has .00 ---
        def po_number_str(val):
            if val is None:
                return ''
            try:
                f = float(val)
                if f.is_integer():
                    return str(int(f))
                return str(val)
            except Exception:
                s = str(val)
                if s.endswith('.00'):
                    return s[:-3]
                return s
        po_no = po_number_str(po_no)
        # Merge PO/budget fields
        dn_row.update({
            'route_type': route_type_val,
            'route_routeLM_metroLM_LMCStandalone': clean_value(po_row.get('route_routeLM_metroLM_LMCStandalone', "")) if po_row else '',
            'ip1_co_built': ip1_co_built_val,
            'dn_recipient': dn_recipient_val,
            'project_name': project_name_val,
            'route_id_site_id': route_id_site_id_val,
            'uid': clean_value(po_row.get('uid', "")) if po_row else '',
            'build_type': build_type_val,
            'category_type': category_type_val,
            # Removed survey_id - no longer needed with one budget per route
            'po_number': po_no,
            'po_length': po_length,
            'parent_route': clean_value(po_row.get('parent_route', "")) if po_row else '',
            'ce_route_lmc_id': '',
            'route_lmc_section_id': '',
            'route_lmc_subsection_id': '',
        })
        # --- PATCH: Budget/Calculated fields ---
        ri_budget_amount_per_meter = ''
        try:
            if budget_row and 'ri_cost_per_meter' in budget_row:
                ri_budget_amount_per_meter = budget_row['ri_cost_per_meter']
            else:
                ri_budget_amount_per_meter = fetch_ri_cost_per_meter_from_supabase(used_site_id)
        except Exception as e:
            pass
        dn_row['ri_budget_amount_per_meter'] = ri_budget_amount_per_meter or ''
        # Calculated fields
        try:
            dn_length = float(dn_row.get('dn_length_mtr') or 0)
            ri_budget = float(dn_row.get('ri_budget_amount_per_meter') or 0)
            actual_total_non_refundable = float(dn_row.get('actual_total_non_refundable') or 0)
            dn_row['projected_budget_ri_amount_dn'] = str(ri_budget * dn_length)
            dn_row['non_refundable_amount_per_mtr'] = str(actual_total_non_refundable / dn_length) if dn_length else ''
            nram = float(dn_row['non_refundable_amount_per_mtr'] or 0)
            dn_row['proj_non_refundable_savings_per_mtr'] = str(ri_budget - nram)
            pbri = float(dn_row['projected_budget_ri_amount_dn'] or 0)
            dn_row['proj_savings_per_dn'] = str(pbri - actual_total_non_refundable)
        except Exception as e:
            dn_row['projected_budget_ri_amount_dn'] = ''
            dn_row['non_refundable_amount_per_mtr'] = ''
            dn_row['proj_non_refundable_savings_per_mtr'] = ''
            dn_row['proj_savings_per_dn'] = ''
        print("📋 FINAL ENRICHED DN DATA:", dn_row)
        return dn_row
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)

        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/validate-parsers")
async def validate_parsers(po_file: UploadFile, dn_file: UploadFile, app_file: UploadFile):
    # ...existing parsing logic...
    po_fields = {}  # Replace with actual extraction logic if needed
    dn_fields = {}  # Replace with actual extraction logic if needed
    app_fields = {}  # Replace with actual extraction logic if needed
    site_id = po_fields.get("SiteID")
    ri_cost_per_meter = fetch_ri_cost_per_meter_from_supabase(site_id)
    # ...other extractions...
    return {
        "po": po_fields,
        "dn": dn_fields,
        "application": app_fields,
        "ri_cost_per_meter_master_budget": ri_cost_per_meter,
    }

NUMERIC_FIELDS = {
    'po_length', 'application_length_mtr', 'dn_length_mtr', 'ot_length', 'dn_ri_amount',
    'multiplying_factor', 'ground_rent', 'administrative_charge', 'supervision_charges',
    'chamber_fee', 'gst', 'ri_budget_amount_per_meter', 'projected_budget_ri_amount_dn',
    'actual_total_non_refundable', 'non_refundable_amount_per_mtr', 'proj_non_refundable_savings_per_mtr',
    'deposit', 'total_dn_amount', 'pit_ri_rate', 'hdd_length',
    'proj_savings_per_dn',
}
INTEGER_FIELDS = {'tat_days', 'no_of_pits'}

def normalize_numeric(val):
    try:
        if val is None or val == '':
            return None
        return float(val)
    except Exception:
        return None

def normalize_integer(val):
    try:
        if val is None or val == '' or (isinstance(val, str) and val.strip() == ''):
            return None
        return int(float(val))
    except Exception:
        return None

@app.post("/api/send-to-master-dn")
async def send_to_master_dn(request: Request):
    print("[LOG] Received request to /api/send-to-master-dn")
    body = await request.json()
    print(f"[LOG] Raw body: {body}")
    data = body.get("data", [])
    print(f"[LOG] Parsed data array: {data}")
    # Build the insert dict using FIELD_MAP to map frontend fields to DB columns
    insert_dict = {}
    for item in data:
        field = item.get("field")
        value = item.get("value")
        print(f"[LOG] Field from frontend: {field}, Value: {value}")
        db_field = FIELD_MAP.get(field, field)  # Map to DB column name
        if db_field not in VALIDATE_PARSER_FIELDS:
            continue
        if db_field in DATE_FIELDS:
            value = normalize_date(value)
        elif db_field in INTEGER_FIELDS:
            value = normalize_integer(value)
        elif db_field in NUMERIC_FIELDS:
            value = normalize_numeric(value)
        insert_dict[db_field] = value
    print(f"[LOG] Final insert_dict (before insert): {insert_dict}")
    # Universal sweep: for any field in insert_dict, if value is '', set to None
    for k, v in insert_dict.items():
        if v == "":
            insert_dict[k] = None
    # Remove sr_no if present, so DB can auto-generate or ignore it
    if 'sr_no' in insert_dict:
        del insert_dict['sr_no']
    for k, v in insert_dict.items():
        print(f"[LOG] Field: {k}, Value: {v}, Type: {type(v)}")
    dn_number = insert_dict.get("dn_number")
    print(f"[LOG] dn_number for duplicate check: {dn_number}")
    if not dn_number:
        print("[ERROR] Missing dn_number in payload.")
        return JSONResponse(status_code=400, content={"error": "Missing dn_number in payload."})
    
    # Ensure dn_number is an integer for duplicate check
    try:
        dn_number_int = int(float(dn_number)) if dn_number else None
        print(f"[LOG] dn_number_int for duplicate check: {dn_number_int}")
    except (ValueError, TypeError):
        print(f"[ERROR] Invalid dn_number format: {dn_number}")
        return JSONResponse(status_code=400, content={"error": "Invalid dn_number format."})
    
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    existing = supabase.table("dn_master").select("dn_number").eq("dn_number", dn_number_int).execute()
    print(f"[LOG] Existing check result: {existing.data}")
    if existing.data and len(existing.data) > 0:
        print(f"[ERROR] DN number {dn_number_int} already exists. Not inserting.")
        return JSONResponse(status_code=409, content={"error": "DN number already exists."})
    try:
        response = supabase.table("dn_master").insert(insert_dict).execute()
        print(f"[LOG] Supabase insert response: {response}")
    except Exception as e:
        print(f"[ERROR] Exception during insert: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})
    print(f"[LOG] Insert successful for dn_number {dn_number_int}")
    return {"success": True}

@app.get("/api/download-master-dn")
def download_master_dn():
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    all_rows = []
    batch_size = 1000
    offset = 0
    while True:
        response = supabase.table("dn_master").select("*").range(offset, offset + batch_size - 1).execute()
        batch = response.data or []
        all_rows.extend(batch)
        print(f"[DEBUG] Fetched batch: {len(batch)} rows (offset {offset})")
        if len(batch) < batch_size:
            break
        offset += batch_size
    print(f"[DEBUG] Downloading {len(all_rows)} rows from dn_master (all batches)")
    data = all_rows
    if not data:
        raise HTTPException(status_code=404, detail="No data found in master DN table.")
    # Convert to DataFrame
    df = pd.DataFrame(data)
    # Clean DataFrame: replace inf/-inf with NA, then fill all NA/NaN with ''
    df = df.replace([float('inf'), float('-inf')], pd.NA)
    df = df.fillna("")
    # Reorder columns to ensure new permit columns are present and in order
    column_order = [
        "route_type", "ip1_co_built", "dn_recipient", "project_name", "route_id_site_id", "uid",
        "build_type", "category_type", "po_number", "po_length", "parent_route",
        "ce_route_lmc_id", "route_lmc_section_id", "route_lmc_subsection_id", "application_number",
        "application_length_mtr", "application_date", "from_location", "to_location", "authority", "ward",
        "dn_number", "dn_length_mtr", "dn_received_date", "trench_type", "ot_length", "surface", "surface_wise_ri_amount",
        "dn_ri_amount", "surface_wise_multiplication_factor", "ground_rent", "administrative_charge", "supervision_charges",
        "chamber_fee", "gst", "ri_budget_amount_per_meter", "projected_budget_ri_amount_dn",
        "actual_total_non_refundable", "non_refundable_amount_per_mtr", "proj_non_refundable_savings_per_mtr",
        "deposit", "total_dn_amount", "new_revised_dn_number", "new_revised_dn_against", "internal_approval_start",
        "internal_approval_end", "ticket_raised_date", "dn_payment_date", "tat_days", "civil_completion_date",
        "hdd_length", "no_of_pits", "pit_ri_rate", "road_name",
        # New permit columns
        "permission_receipt_date", "permit_no", "permit_start_date", "permit_end_date", "permitted_length_by_ward_mts"
    ]
    # Add any missing columns to the DataFrame (if not present)
    for col in column_order:
        if col not in df.columns:
            df[col] = ""
    df = df.reindex(columns=column_order)
    # Write to a temp Excel file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        excel_path = tmp.name
    # Use ExcelWriter for formatting
    with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
        # Write data without header, start at row 1
        df.to_excel(writer, index=False, sheet_name='MasterDN', header=False, startrow=1)
        workbook = writer.book
        worksheet = writer.sheets['MasterDN']

        # Header format: light blue, bold, centered, wrapped, border
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#B7E1FC',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True
        })

        # Data cell format: centered, wrapped, border
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True
        })

        # Write headers with format
        worksheet.set_row(0, 38)
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)

        # Write data cells with format and set row height
        for row in range(df.shape[0]):
            worksheet.set_row(row + 1, 28)
            for col in range(df.shape[1]):
                worksheet.write(row + 1, col, df.iloc[row, col], cell_format)

        # Set column widths for readability (no format here)
        for i, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            worksheet.set_column(i, i, min(max_len + 2, 40))

        worksheet.freeze_panes(1, 0)
    # Return as file download
    return FileResponse(excel_path, filename="Master_DN_Database.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.post("/api/upload-dn-master")
async def upload_dn_master(file: UploadFile = File(...)):
    start_total = time.time()
    # 1. Read Excel file into DataFrame
    start_read = time.time()
    contents = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")
    print(f"[TIMING] Read Excel: {time.time() - start_read:.3f} seconds")

    # 2. Define your required DB columns (should match your dn_master schema)
    required_columns = [
        "route_type", "ip1_co_built", "dn_recipient", "project_name", "route_id_site_id", "uid",
        "build_type", "category_type", "po_number", "po_length", "parent_route",
        "ce_route_lmc_id", "route_lmc_section_id", "route_lmc_subsection_id", "application_number",
        "application_length_mtr", "application_date", "from_location", "to_location", "authority", "ward",
        "dn_number", "dn_length_mtr", "dn_received_date", "trench_type", "ot_length", "surface", "surface_wise_ri_amount",
        "dn_ri_amount", "surface_wise_multiplication_factor", "ground_rent", "administrative_charge", "supervision_charges",
        "chamber_fee", "gst", "ri_budget_amount_per_meter", "projected_budget_ri_amount_dn",
        "actual_total_non_refundable", "non_refundable_amount_per_mtr", "proj_non_refundable_savings_per_mtr",
        "deposit", "total_dn_amount", "new_revised_dn_number", "new_revised_dn_against", "internal_approval_start",
        "internal_approval_end", "ticket_raised_date", "dn_payment_date", "tat_days", "civil_completion_date",
        "hdd_length", "no_of_pits", "pit_ri_rate", "road_name",
        # New permit columns
        "permission_receipt_date", "permit_no", "permit_start_date", "permit_end_date", "permitted_length_by_ward_mts"
    ]

    # 3. Validate columns
    missing = [col for col in required_columns if col not in df.columns]
    extra = [col for col in df.columns if col not in required_columns]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {missing}")
    # Optionally, warn about extra columns

    # 4. Clean DataFrame (replace NaN/inf)
    start_clean = time.time()
    df = df.replace([float('inf'), float('-inf')], pd.NA)
    df = df.fillna("")
    NUMERIC_FIELDS = {
        'po_length', 'application_length_mtr', 'dn_length_mtr', 'ot_length', 'dn_ri_amount',
        'multiplying_factor', 'ground_rent', 'administrative_charge', 'supervision_charges',
        'chamber_fee', 'gst', 'ri_budget_amount_per_meter', 'projected_budget_ri_amount_dn',
        'actual_total_non_refundable', 'non_refundable_amount_per_mtr', 'proj_non_refundable_savings_per_mtr',
        'deposit', 'total_dn_amount', 'pit_ri_rate', 'hdd_length',
        'proj_savings_per_dn',
    }
    INTEGER_FIELDS = {'tat_days', 'no_of_pits'}
    DATE_FIELDS = {
        'application_date',
        'dn_received_date',
        'internal_approval_start',
        'internal_approval_end',
        'ticket_raised_date',
        'dn_payment_date',
        'civil_completion_date'
    }
    def normalize_numeric(val):
        try:
            if val is None or val == '':
                return None
            return float(val)
        except Exception:
            return None
    def normalize_integer(val):
        try:
            if val is None or val == '' or (isinstance(val, str) and val.strip() == ''):
                return None
            return int(float(val))
        except Exception:
            return None
    def normalize_date(val):
        if val is None or val == "":
            return None
        if isinstance(val, str):
            # Try DD/MM/YYYY
            try:
                return datetime.strptime(val, '%d/%m/%Y').strftime('%Y-%m-%d')
            except Exception:
                pass
            # Try YYYY-MM-DD (already correct)
            try:
                return datetime.strptime(val, '%Y-%m-%d').strftime('%Y-%m-%d')
            except Exception:
                pass
        return None  # Return None if not a recognized date string
    cleaned_rows = []
    for idx, row in df.iterrows():
        data = row.to_dict()
        # Normalize all fields before upsert
        for k, v in data.items():
            if k in NUMERIC_FIELDS:
                data[k] = normalize_numeric(v)
            elif k in INTEGER_FIELDS:
                data[k] = normalize_integer(v)
            elif k in DATE_FIELDS:
                data[k] = normalize_date(v)
            elif v == "":
                data[k] = None
        dn_number = data.get("dn_number")
        if not dn_number:
            continue
        cleaned_rows.append(data)
    print(f"[TIMING] Clean/Map Rows: {time.time() - start_clean:.3f} seconds")
    print("[DN MASTER CLEANED ROWS]", cleaned_rows[:3])

    # 5. Bulk upsert all rows (using dn_number as unique key)
    start_upsert = time.time()
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    errors = []
    try:
        response = supabase.table("dn_master").upsert(cleaned_rows, on_conflict="dn_number").execute()
        if hasattr(response, 'error') and response.error:
            errors.append(str(response.error))
    except Exception as e:
        errors.append(str(e))
    print(f"[TIMING] Upsert to Supabase: {time.time() - start_upsert:.3f} seconds")
    print(f"[TIMING] Total /api/upload-dn-master: {time.time() - start_total:.3f} seconds")
    if errors:
        return {"success": False, "errors": errors}
    return {"success": True, "message": "All rows upserted successfully."}

@app.post("/api/upload-budget-master")
@app.post("/api/fullroute-upload-master")
async def upload_budget_master(file: UploadFile = File(...)):
    import time
    start_total = time.time()
    # 1. Read Excel file into DataFrame
    start_read = time.time()
    contents = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")
    print(f"[TIMING] Read Excel: {time.time() - start_read:.3f} seconds")

    # 2. Define the required schema columns (should match your budget_master schema)
    schema_columns = [
        "id",
        "route_id_site_id",
        "ce_length_mtr",
        "ri_cost_per_meter",
        "material_cost_per_meter",
        "build_cost_per_meter",
        "total_ri_amount",
        "material_cost",
        "execution_cost_including_hh",
        "total_cost_without_deposit",
        "route_type",
    ]
    numeric_columns = {
        "ce_length_mtr",
        "ri_cost_per_meter",
        "material_cost_per_meter",
        "build_cost_per_meter",
        "total_ri_amount",
        "material_cost",
        "execution_cost_including_hh",
        "total_cost_without_deposit",
    }

    # 3. Clean and map rows
    start_clean = time.time()
    cleaned_rows = []
    for idx, row in df.iterrows():
        cleaned = {}
        for col in schema_columns:
            value = row[col] if col in row else None
            if pd.isna(value) or value == "":
                value = None
            if col in numeric_columns and value is not None:
                try:
                    value = float(value)
                except Exception:
                    value = None
            cleaned[col] = value
        # Remove id if present (autoincrement)
        if "id" in cleaned:
            cleaned.pop("id")
        cleaned_rows.append(cleaned)
    print(f"[TIMING] Clean/Map Rows: {time.time() - start_clean:.3f} seconds")
    print("[BUDGET MASTER CLEANED ROWS]", cleaned_rows[:3])

    # 4. Complete replacement: delete all existing records then insert new ones
    start_upsert = time.time()
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    errors = []
    try:
        # First, delete all existing records in budget_master
        print("[BUDGET MASTER] Deleting all existing records...")
        delete_response = supabase.table("budget_master").delete().neq("id", 0).execute()
        print(f"[BUDGET MASTER] Deleted existing records, response: {delete_response}")
        
        # Then insert all new rows from Excel
        print(f"[BUDGET MASTER] Inserting {len(cleaned_rows)} new records...")
        response = supabase.table("budget_master").insert(cleaned_rows).execute()
        print(f"[BUDGET MASTER] Insert response: {response}")
        if hasattr(response, 'error') and response.error:
            errors.append(str(response.error))
    except Exception as e:
        print(f"[BUDGET MASTER] Exception during complete replacement: {e}")
        errors.append(str(e))
    print(f"[TIMING] Complete replacement to Supabase: {time.time() - start_upsert:.3f} seconds")
    print(f"[TIMING] Total /api/upload-budget-master: {time.time() - start_total:.3f} seconds")
    return {
        "success": len(errors) == 0,
        "errors": errors,
        "rows": len(cleaned_rows),
        "cleaned_rows": cleaned_rows[:5],
        "message": "All rows upserted successfully." if len(errors) == 0 else "Some rows failed."
    }

@app.get("/api/download-master-budget")
def download_master_budget():
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    all_rows = []
    batch_size = 1000
    offset = 0
    while True:
        response = supabase.table("budget_master").select("*").range(offset, offset + batch_size - 1).execute()
        batch = response.data or []
        all_rows.extend(batch)
        print(f"[DEBUG] Fetched batch: {len(batch)} rows (offset {offset})")
        if len(batch) < batch_size:
            break
        offset += batch_size
    print(f"[DEBUG] Downloading {len(all_rows)} rows from budget_master (all batches)")
    data = all_rows
    if not data:
        raise HTTPException(status_code=404, detail="No data found in budget_master table.")
    # Use the column order from the frontend
    supabase_headers = [
        "id",
        "route_id_site_id",
        "ce_length_mtr",
        "ri_cost_per_meter",
        "material_cost_per_meter",
        "build_cost_per_meter",
        "total_ri_amount",
        "material_cost",
        "execution_cost_including_hh",
        "total_cost_without_deposit",
        "route_type",
    ]
    # Convert to DataFrame and reorder columns
    df = pd.DataFrame(data)
    df = df.reindex(columns=supabase_headers)
    df = df.replace([float('inf'), float('-inf')], pd.NA)
    df = df.fillna("")
    # Write to a temp Excel file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        excel_path = tmp.name
    with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='MasterBudget', header=False, startrow=1)
        workbook = writer.book
        worksheet = writer.sheets['MasterBudget']
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#B7E1FC',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True
        })
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True
        })
        worksheet.set_row(0, 38)
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
        for row in range(df.shape[0]):
            worksheet.set_row(row + 1, 28)
            for col in range(df.shape[1]):
                worksheet.write(row + 1, col, df.iloc[row, col], cell_format)
        for i, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            worksheet.set_column(i, i, min(max_len + 2, 40))
        worksheet.freeze_panes(1, 0)
    return FileResponse(excel_path, filename="Master_Budget_Database.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.post("/api/upload-po-master")
async def upload_po_master(file: UploadFile = File(...)):
    start_total = time.time()
    # 1. Read Excel file into DataFrame
    start_read = time.time()
    contents = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")
    print("[PO MASTER] DataFrame dtypes after read_excel:", df.dtypes)
    print("[PO MASTER] DataFrame head after read_excel:\n", df.head())

    # 2. Define the required schema columns
    schema_columns = [
        "route_id_site_id",
        "parent_route",
        "route_type",
        "uid",
        "po_no_cobuild",
        "po_length_cobuild",
        "po_no_ip1",
        "po_length_ip1",
        "route_routeLM_metroLM_LMCStandalone",
    ]
    numeric_columns = {"po_length_cobuild", "po_length_ip1"}

    # 3. Normalize and map columns from Excel to schema
    start_clean = time.time()
    def normalize(col):
        return str(col).strip().lower().replace(" ", "_").replace("/", "_")
    print("[PO MASTER] Raw Excel headers:", list(df.columns))
    print("[PO MASTER] Normalized Excel headers:", [normalize(col) for col in df.columns])
    print("[PO MASTER] schema_columns:", schema_columns)
    excel_col_map = {normalize(col): col for col in df.columns}
    schema_col_map = {normalize(col): col for col in schema_columns}
    col_mapping = {}
    for norm_col, excel_col in excel_col_map.items():
        if norm_col in schema_col_map:
            col_mapping[schema_col_map[norm_col]] = excel_col
    if 'route_id_site_id' not in col_mapping:
        for col in df.columns:
            if normalize(col) in ['route_id_site_id', 'route_id__site_id']:
                col_mapping['route_id_site_id'] = col
                break
    print("[PO MASTER] Column mapping:", col_mapping)
    cleaned_rows = []
    for idx, row in df.iterrows():
        cleaned = {}
        for schema_col in schema_columns:
            excel_col = col_mapping.get(schema_col)
            value = row[excel_col] if excel_col in row else None
            if pd.isna(value) or value == "":
                value = None
            if schema_col in numeric_columns and value is not None:
                try:
                    value = float(value)
                except Exception:
                    value = None
            # Force PO numbers to string, even if they look like numbers
            if schema_col in ['po_no_ip1', 'po_no_cobuild'] and value is not None:
                if isinstance(value, float) and value.is_integer():
                    value = str(int(value))
                else:
                    value = str(value)
            cleaned[schema_col] = value
        print(f"[PO MASTER] Cleaned row {idx}:", cleaned)
        cleaned_rows.append(cleaned)
    print(f"[PO MASTER] Clean/Map Rows: {time.time() - start_clean:.3f} seconds")
    print("[PO MASTER] Final cleaned_rows sample:", cleaned_rows[:3])
    print("[PO MASTER] Col mapping:", col_mapping)

    # 5. Upsert all rows in bulk (using route_id_site_id as unique key)
    start_upsert = time.time()
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    errors = []
    print("[PO MASTER] About to upsert to Supabase")
    try:
        response = supabase.table("po_master").upsert(cleaned_rows, on_conflict="route_id_site_id").execute()
        print("[PO MASTER] Upsert response:", response)
        if hasattr(response, 'error') and response.error:
            errors.append(str(response.error))
    except Exception as e:
        print("[PO MASTER] Exception during upsert:", e)
        errors.append(str(e))
    print("[PO MASTER] Finished upsert block")
    print(f"[TIMING] Upsert to Supabase: {time.time() - start_upsert:.3f} seconds")
    print(f"[TIMING] Total /api/upload-po-master: {time.time() - start_total:.3f} seconds")
    return {
        "success": len(errors) == 0,
        "errors": errors,
        "rows": len(cleaned_rows),
        "cleaned_rows": cleaned_rows[:5],
        "col_mapping": col_mapping,
        "message": "All rows upserted successfully." if len(errors) == 0 else "Some rows failed."
    }

@app.get("/api/download-master-po")
def download_master_po():
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    all_rows = []
    batch_size = 1000
    offset = 0
    while True:
        response = supabase.table("po_master").select("*").range(offset, offset + batch_size - 1).execute()
        batch = response.data or []
        all_rows.extend(batch)
        print(f"[DEBUG] Fetched batch: {len(batch)} rows (offset {offset})")
        if len(batch) < batch_size:
            break
        offset += batch_size
    print(f"[DEBUG] Downloading {len(all_rows)} rows from po_master (all batches)")
    data = all_rows
    if not data:
        raise HTTPException(status_code=404, detail="No data found in po_master table.")
    # Use the column order from the schema
    po_headers = [
        "route_id_site_id",
        "parent_route",
        "route_type",
        "uid",
        "po_no_cobuild",
        "po_length_cobuild",
        "po_no_ip1",
        "po_length_ip1",
        "route_routeLM_metroLM_LMCStandalone",
    ]
    # Convert to DataFrame and reorder columns
    df = pd.DataFrame(data)
    df = df.reindex(columns=po_headers)
    df = df.replace([float('inf'), float('-inf')], pd.NA)
    df = df.fillna("")
    # Write to a temp Excel file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        excel_path = tmp.name
    with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='MasterPO', header=False, startrow=1)
        workbook = writer.book
        worksheet = writer.sheets['MasterPO']
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#B7E1FC',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True
        })
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True
        })
        worksheet.set_row(0, 38)
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
        for row in range(df.shape[0]):
            worksheet.set_row(row + 1, 28)
            for col in range(df.shape[1]):
                worksheet.write(row + 1, col, df.iloc[row, col], cell_format)
        for i, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            worksheet.set_column(i, i, min(max_len + 2, 40))
        worksheet.freeze_panes(1, 0)
    return FileResponse(excel_path, filename="Master_PO_Database.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.post("/api/clean-excel")
async def clean_excel(file: UploadFile = File(...)):
    # Save uploaded file to a temp location
    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[-1]) as tmp_in:
        tmp_in.write(await file.read())
        tmp_in_path = tmp_in.name
    # Prepare output path
    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[-1]) as tmp_out:
        tmp_out_path = tmp_out.name
    # Call the cleaning function (for now, just copy input to output)
    # Replace this with your real cleaning logic
    result = clean_excel_file(tmp_in_path, tmp_out_path)
    # Return the cleaned file for download
    return FileResponse(tmp_out_path, filename=f"cleaned_{file.filename}", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

from fastapi import UploadFile, File
from fastapi.responses import JSONResponse
import tempfile
import os
from pdf2image import convert_from_path
import base64
# --- NMMC Extraction Endpoint ---
@app.post("/api/nmmc-extract")
async def nmmc_extract(file: UploadFile = File(...)):
    import tempfile, os
    from parsers import nmmc
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = os.path.join(temp_dir, file.filename)
        with open(temp_path, "wb") as tmp:
            tmp.write(await file.read())
        # Use the new comprehensive parser with hardcoded values
        non_refundable_result = nmmc.non_refundable_request_parser(temp_path)
        # Get SD data using the AI result from non_refundable parser
        all_fields = nmmc.extract_nmmc_all_fields(temp_path)
        sd_headers, sd_row = nmmc.sd_parser_from_ai_result(all_fields)
        
        # Debug output
        print(f"[NMMC DEBUG] non_refundable_result keys: {list(non_refundable_result.keys())}")
        print(f"[NMMC DEBUG] Sample non_refundable values:")
        for key, value in list(non_refundable_result.items())[:5]:
            print(f"  {key}: {value}")
        
        return {
            "non_refundable": non_refundable_result,
            "sd_headers": sd_headers,
            "sd_row": sd_row,
            "all_fields": all_fields
        }

def extract_demand_note_number(row: dict, type_: str) -> str:
    keys = (
        ["Demand Note Reference number", "Demand Note No.", "DN No", "Demand Note Number"]
        if type_ == "non_refundable" else
        ["DN No", "Demand Note Reference number", "Demand Note No.", "Demand Note Number"]
    )
    for key in keys:
        if key in row and str(row[key]).strip():
            return str(row[key]).strip().replace("/", "_").replace(" ", "_")
    return "Output"


def style_excel(ws, headers):
    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    # Header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 30
    # Data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = center
            cell.border = border

@app.post("/api/excel/non_refundable")
async def excel_non_refundable(
    payload: Dict = Body(...)
):
    rows: List[Dict] = payload.get("rows", [])
    if not rows:
        return {"error": "No data provided"}
    headers = list(rows[0].keys())
    wb = Workbook()
    ws = wb.active
    ws.title = "Non-Refundable"
    # Write header
    ws.append(headers)
    # Write data
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_excel(ws, headers)
    # Filename
    dn_num = extract_demand_note_number(rows[0], "non_refundable")
    filename = f"{dn_num}_Non_Refundable_Output.xlsx"
    temp_path = os.path.join(tempfile.gettempdir(), f"{uuid.uuid4()}_{filename}")
    wb.save(temp_path)
    return FileResponse(temp_path, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.post("/api/excel/sd")
async def excel_sd(
    payload: Dict = Body(...)
):
    rows: List[Dict] = payload.get("rows", [])
    if not rows:
        return {"error": "No data provided"}
    headers = list(rows[0].keys())
    wb = Workbook()
    ws = wb.active
    ws.title = "SD"
    # Write header
    ws.append(headers)
    # Write data
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_excel(ws, headers)
    # Filename
    dn_num = extract_demand_note_number(rows[0], "sd")
    filename = f"{dn_num}_SD_Output.xlsx"
    temp_path = os.path.join(tempfile.gettempdir(), f"{uuid.uuid4()}_{filename}")
    wb.save(temp_path)
    return FileResponse(temp_path, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.post("/api/extract-pdf-debug")
async def extract_pdf_debug(pdf_file: UploadFile = File(...)):
    import fitz  # PyMuPDF
    import camelot
    import tempfile, os
    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, f"debug_{pdf_file.filename}")
    file_bytes = await pdf_file.read()
    with open(temp_path, "wb") as f:
        f.write(file_bytes)
    try:
        # Extract text
        doc = fitz.open(temp_path)
        text = "\n".join(page.get_text() for page in doc)
        doc.close()
        # Extract tables
        tables = []
        try:
            camelot_tables = camelot.read_pdf(temp_path, pages='all', flavor='stream')
            for table in camelot_tables:
                tables.append(table.df.values.tolist())
        except Exception as e:
            tables = [[f"Camelot extraction failed: {e}"]]
        # Also print robust universal application parser result
        from parsers.universal_application_parser import universal_application_parser
        parsed = universal_application_parser(temp_path)
        print("📋 EXTRACT PDF DEBUG RESULT:", parsed)
        os.remove(temp_path)
        return {"status": "done"}
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        print(f"Error extracting PDF: {e}")
        return {"error": str(e)}

@app.post("/api/nmmc-debug-extract")
async def nmmc_debug_extract(dn_file: UploadFile = File(...)):
    """
    Debug endpoint for NMMC DN parsing - prints full OCR text, best possible tables, and robustly extracted fields.
    """
    import tempfile
    import os
    from pdf2image import convert_from_path
    import numpy as np
    import pytesseract
    from PIL import Image
    import camelot
    import cv2
    from parsers.nmmc import extract_nmmc_fields_opencv_ocr
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = os.path.join(temp_dir, f"nmmc_debug_{dn_file.filename}")
            with open(temp_path, "wb") as f:
                f.write(await dn_file.read())
            pages = convert_from_path(temp_path, dpi=300)
            # Try table extraction 
            found_any_table = False
            for flavor in ['lattice', 'stream']:
                try:
                    tables = camelot.read_pdf(temp_path, pages='all', flavor=flavor)
                    if len(tables) > 0:
                        found_any_table = True
                        break  # Prefer first mode with tables
                except Exception as e:
                    pass
            if not found_any_table:
                print("========== OPENCV+TESSERACT TABLE EXTRACTION (IMAGE-BASED) ==========")
                for page_num, pil_img in enumerate(pages):
                    print(f"\n--- PAGE {page_num + 1} ---")
                    img = np.array(pil_img.convert('L'))
                    _, bin_img = cv2.threshold(img, 180, 255, cv2.THRESH_BINARY_INV)
                    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40,1))
                    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1,40))
                    horizontal_lines = cv2.morphologyEx(bin_img, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
                    vertical_lines = cv2.morphologyEx(bin_img, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
                    table_mask = cv2.add(horizontal_lines, vertical_lines)
                    contours, _ = cv2.findContours(table_mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
                    bounding_boxes = [cv2.boundingRect(c) for c in contours]
                    bounding_boxes = sorted(bounding_boxes, key=lambda b: (b[1], b[0]))
                    cells = [b for b in bounding_boxes if b[2] > 40 and b[3] > 15]
                    # Group cells into rows (use 35px threshold)
                    def block_to_rows(block):
                        block = sorted(block, key=lambda b: (b[1], b[0]))
                        rows = []
                        current_row = []
                        last_y = None
                        for b in block:
                            x, y, w, h = b
                            if last_y is None or abs(y - last_y) < 35:
                                current_row.append(b)
                                last_y = y
                            else:
                                rows.append(sorted(current_row, key=lambda b: b[0]))
                                current_row = [b]
                                last_y = y
                        if current_row:
                            rows.append(sorted(current_row, key=lambda b: b[0]))
                        return rows
                    if cells:
                        rows = block_to_rows(cells)
                        table_grid = []
                        for row in rows:
                            row_text = []
                            for x, y, w, h in row:
                                cell_img = img[y:y+h, x:x+w]
                                cell_text = pytesseract.image_to_string(cell_img, config='--psm 7', lang='mar+eng').strip().replace('\n', ' ')
                                row_text.append(cell_text if cell_text else '')
                            table_grid.append(row_text)
                        if table_grid:
                            max_cols = max(len(r) for r in table_grid)
                            for r in table_grid:
                                r += [''] * (max_cols - len(r))
                            headers = [f"Col {i+1}" for i in range(max_cols)]
                            from tabulate import tabulate
                            print(tabulate(table_grid, headers=headers, showindex="always", tablefmt="grid", stralign="center"))
                        else:
                            print("[No table-like structure detected on this page]")
                    else:
                        print("[No table-like structure detected on this page]")
            print("========== END TABLE EXTRACTION ==========")
            # 3. Print robust extracted fields
            result = extract_nmmc_fields_opencv_ocr(temp_path)
            print("========== ROBUST EXTRACTED FIELDS ==========")
            for k, v in result.items():
                print(f"{k}: {v}")
            print("========== NMMC DEBUG EXTRACTION END ==========")
            return {"extracted_fields": result}
    except Exception as e:
        print(f"Error in NMMC debug extraction: {e}")
        return {"error": str(e)}

@app.post("/api/debug-mcgm-application")
async def debug_mcgm_application(file: UploadFile = File(...)):
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    temp.write(await file.read())
    temp.close()
    try:
        # Only use the universal parser's extract_road_name
        from parsers.universal_application_parser import extract_road_name
        doc = fitz.open(temp.name)
        text = "\n".join(page.get_text() for page in doc)
        doc.close()
        road_name = extract_road_name(text)
        print("📋 MCGM APPLICATION DEBUG - Road Name:", road_name)
    except Exception as e:
        os.remove(temp.name)
        print(f"[ERROR] {e}")
        return {"success": False, "error": str(e)}
    os.remove(temp.name)
    return {"success": True}

@app.get("/api/generate-master-tracker")
def generate_master_tracker():
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    all_rows = []
    batch_size = 1000
    offset = 0
    while True:
        response = supabase.table("dn_master").select("*").range(offset, offset + batch_size - 1).execute()
        batch = response.data or []
        all_rows.extend(batch)
        if len(batch) < batch_size:
            break
        offset += batch_size
    print(f"📋 MASTER TRACKER: Generated {len(all_rows)} rows")
    data = all_rows
    if not data:
        raise HTTPException(status_code=404, detail="No data found in dn_master table.")

    # Define the master tracker columns and their sources (snake_case, matching DB)
    columns = [
        ("circle", lambda row: "Mumbai"),
        ("state", lambda row: "Maharashtra"),
        ("city", lambda row: "Mumbai"),
        ("area", lambda row: row.get("ward", "")),
        ("type_of_build", lambda row: row.get("ip1_co_built", "")),
        ("financial_year", lambda row: _get_financial_year(row.get("dn_received_date", ""))),
        ("link_name", lambda row: row.get("route_id_site_id", "")),
        ("ce_route_id", lambda row: row.get("ce_route_lmc_id", "")),
        ("ce_section_id", lambda row: row.get("route_lmc_section_id", "")),
        ("ce_subsection_id", lambda row: row.get("route_lmc_subsection_id", "")),
        ("municipal_authority", lambda row: row.get("authority", "")),
        ("ward", lambda row: row.get("ward", "")),
        ("road_name", lambda row: row.get("road_name", "")),
        ("start_point", lambda row: row.get("from_location", "")),
        ("end_point", lambda row: row.get("to_location", "")),
        ("authority_application_date", lambda row: row.get("application_date", "")),
        ("authority_application_number", lambda row: row.get("application_number", "")),
        ("authority_new_application_number", lambda row: ""),
        ("dn_received_date", lambda row: row.get("dn_received_date", "")),
        ("dn_number", lambda row: row.get("dn_number", "")),
        ("dn_length", lambda row: row.get("dn_length_mtr", "")),
        ("ri_amount", lambda row: row.get("dn_ri_amount", "")),
        ("sd_amount", lambda row: row.get("deposit", "")),
        ("supervision_charge_amt", lambda row: row.get("supervision_charges", "")),
        ("additional_charge_amt", lambda row: ""),
        ("land_length_mtrs", lambda row: row.get("dn_length_mtr", "")),
        ("land_rent_start_date", lambda row: ""),
        ("land_rent_end_date", lambda row: ""),
        ("land_rent_tenure", lambda row: ""),
        ("ground_rent", lambda row: row.get("ground_rent", "")),
        ("access_charge", lambda row: ""),
        ("admin_charges", lambda row: row.get("administrative_charge", "")),
        ("chamber_fee", lambda row: row.get("chamber_fee", "")),
        ("gst", lambda row: row.get("gst", "")),
        ("dn_amount_without_gst", lambda row: _calc_without_gst(row)),
        ("dn_amount_with_gst", lambda row: row.get("total_dn_amount", "")),
        ("dn_payment_date", lambda row: row.get("dn_payment_date", "")),
        ("differential_dn_number", lambda row: ""),
        ("differential_dn_amount", lambda row: ""),
        ("differential_dn_sd", lambda row: ""),
        ("permission_receipt_date", lambda row: row.get("permission_receipt_date", "")),
        ("permit_no", lambda row: row.get("permit_no", "")),
        ("permit_start_date", lambda row: row.get("permit_start_date", "")),
        ("permit_end_date", lambda row: row.get("permit_end_date", "")),
        ("permitted_length_by_ward_mts", lambda row: row.get("permitted_length_by_ward_mts", "")),
        ("permitted_no_of_ducts", lambda row: ""),
        ("work_status", lambda row: ""),
        ("date_of_refund_application", lambda row: ""),
        ("refund_application_no", lambda row: ""),
        ("sd_refund_amount", lambda row: ""),
        ("sd_deductions", lambda row: ""),
        ("remarks_for_deductions", lambda row: ""),
        ("sd_check", lambda row: ""),
        ("sd_refund_date", lambda row: ""),
        ("sd_refund_ack_no", lambda row: ""),
        ("date_of_refund_application_ri", lambda row: ""),
        ("refund_application_no_ri", lambda row: ""),
        ("ri_refund_amount", lambda row: ""),
        ("penalty_deductions", lambda row: ""),
        ("remarks_for_deductions_ri", lambda row: ""),
        ("ri_refund_date", lambda row: ""),
        ("ri_refund_ack_no", lambda row: ""),
    ]

    # List of manual fields (not overwritten unless blank)
    MANUAL_FIELDS = [
        "authority_new_application_number", "additional_charge_amt", "land_rent_start_date", "land_rent_end_date", "land_rent_tenure", "access_charge", "differential_dn_number", "differential_dn_amount", "differential_dn_sd", "permitted_no_of_ducts", "work_status", "date_of_refund_application", "refund_application_no", "sd_refund_amount", "sd_deductions", "remarks_for_deductions", "sd_check", "sd_refund_date", "sd_refund_ack_no", "date_of_refund_application_ri", "refund_application_no_ri", "ri_refund_amount", "penalty_deductions", "remarks_for_deductions_ri", "ri_refund_date", "ri_refund_ack_no"
    ]

    # Fetch all existing rows from master_tracker
    existing = {row['dn_number']: row for row in supabase.table("master_tracker").select("*").execute().data or []}

    TEXT_FIELDS = {
        "dn_number", "circle", "state", "city", "area", "type_of_build", "financial_year", "link_name",
        "ce_route_id", "ce_section_id", "ce_subsection_id", "municipal_authority", "ward", "road_name",
        "start_point", "end_point", "authority_application_number", "authority_new_application_number",
        "differential_dn_number", "permit_no", "permitted_no_of_ducts", "work_status", "refund_application_no",
        "remarks_for_deductions", "sd_check", "sd_refund_ack_no", "refund_application_no_ri",
        "remarks_for_deductions_ri", "ri_refund_ack_no"
    }
    # Assemble the data rows and upsert dicts
    rows = []
    upsert_dicts = []
    for row in data:
        row_values = [func(row) for _, func in columns]
        upsert_dict = {col: val for (col, _), val in zip(columns, row_values)}
        dn_number = upsert_dict["dn_number"]
        # For manual fields, preserve existing value if present
        if dn_number in existing:
            for field in MANUAL_FIELDS:
                existing_val = existing[dn_number].get(field)
                if existing_val not in [None, "", "null"]:
                    upsert_dict[field] = existing_val
        # Convert "" to None for all non-text fields
        for k in list(upsert_dict.keys()):
            if upsert_dict[k] == "" and k not in TEXT_FIELDS:
                upsert_dict[k] = None
        upsert_dicts.append(upsert_dict)
        rows.append(row_values)

    # Upsert all rows into master_tracker
    if upsert_dicts:
        supabase.table("master_tracker").upsert(upsert_dicts, on_conflict="dn_number").execute()

    # Fetch all rows from master_tracker for export
    tracker_resp = supabase.table("master_tracker").select("*").execute()
    tracker_data = tracker_resp.data or []
    # Rebuild rows for export in the same order/format
    export_rows = []
    for row in tracker_data:
        export_rows.append([row.get(col, "") for col, _ in columns])
    df = pd.DataFrame(export_rows, columns=[col for col, _ in columns])

    # Replace NaN, inf, -inf with empty string before Excel export
    import numpy as np
    df = df.replace([np.inf, -np.inf], np.nan)
    df = df.fillna("")

    # Write to Excel in memory (formatting as before)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='MasterTracker')
        workbook = writer.book
        worksheet = writer.sheets['MasterTracker']
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#B7E1FC',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True
        })
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True
        })
        worksheet.set_row(0, 38)
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
        for row in range(df.shape[0]):
            worksheet.set_row(row + 1, 28)
            for col in range(df.shape[1]):
                worksheet.write(row + 1, col, df.iloc[row, col], cell_format)
        for i, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(str(col)))
            worksheet.set_column(i, i, min(max_len + 2, 40))
        worksheet.freeze_panes(1, 0)
    output.seek(0)
    return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": "attachment; filename=MasterTracker.xlsx"})

@app.post("/api/parse-permit")
async def parse_permit_file(file: UploadFile = File(...)):
    import tempfile, os
    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, f"permit_{file.filename}")
    file_bytes = await file.read()
    with open(temp_path, "wb") as f:
        f.write(file_bytes)
    try:
        result = extract_permit_fields(temp_path)
        os.remove(temp_path)
        return result
    except Exception as e:
        os.remove(temp_path)
        return {"error": str(e)}

@app.post("/api/upsert-permit-fields")
async def upsert_permit_fields_api(request: Request):
    try:
        permit_fields = await request.json()
        result = upsert_permit_fields_to_dn_master(permit_fields)
        if result and hasattr(result, 'error') and result.error:
            return JSONResponse(status_code=500, content={"error": str(result.error)})
        return {"success": True}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

def _get_financial_year(date_str):
    # date_str format: YYYY-MM-DD or DD/MM/YYYY
    try:
        if not date_str:
            return ""
        if "-" in date_str:
            dt = pd.to_datetime(date_str, errors='coerce')
        else:
            dt = pd.to_datetime(date_str, dayfirst=True, errors='coerce')
        if pd.isnull(dt):
            return ""
        year = dt.year
        month = dt.month
        if month >= 4:
            fy_start = year
            fy_end = year + 1
        else:
            fy_start = year - 1
            fy_end = year
        return f"{str(fy_start)[-2:]}-{str(fy_end)[-2:]}"
    except Exception:
        return ""

def _calc_without_gst(row):
    try:
        total = float(row.get("total_dn_amount", 0) or 0)
        gst = float(row.get("gst", 0) or 0)
        return str(total - gst)
    except Exception:
        return ""

# Add this utility near the top (after imports)
def clean_json(obj):
    import math
    import numpy as np
    if isinstance(obj, dict):
        return {k: clean_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_json(x) for x in obj]
    elif isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    elif isinstance(obj, (np.floating, np.float32, np.float64)):
        if np.isnan(obj) or np.isinf(obj):
            return None
        return float(obj)
    elif isinstance(obj, (np.integer,)):
        return int(obj)
    else:
        return obj

@app.post("/api/upload-master-tracker")
async def upload_master_tracker(file: UploadFile = File(...)):
    import pandas as pd
    import numpy as np
    import math
    import logging
    contents = await file.read()
    print("[UPLOAD MASTER TRACKER] Received file of size:", len(contents))
    try:
        df = pd.read_excel(io.BytesIO(contents))
        print("[UPLOAD MASTER TRACKER] DataFrame head:\n", df.head())
        print("[UPLOAD MASTER TRACKER] DataFrame columns:", list(df.columns))
    except Exception as e:
        print("[UPLOAD MASTER TRACKER] Error reading Excel:", e)
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    # Print all columns expected by the DB
    expected_columns = [
        "dn_number", "circle", "state", "city", "area", "type_of_build", "financial_year", "link_name", "ce_route_id", "ce_section_id", "ce_subsection_id", "municipal_authority", "ward", "road_name", "start_point", "end_point", "authority_application_date", "authority_application_number", "authority_new_application_number", "dn_received_date", "dn_length", "ri_amount", "sd_amount", "supervision_charge_amt", "additional_charge_amt", "land_length_mtrs", "land_rent_start_date", "land_rent_end_date", "land_rent_tenure", "ground_rent", "access_charge", "admin_charges", "chamber_fee", "gst", "dn_amount_without_gst", "dn_amount_with_gst", "dn_payment_date", "differential_dn_number", "differential_dn_amount", "differential_dn_sd", "permission_receipt_date", "permit_no", "permit_start_date", "permit_end_date", "permitted_length_by_ward_mts", "permitted_no_of_ducts", "work_status", "date_of_refund_application", "refund_application_no", "sd_refund_amount", "sd_deductions", "remarks_for_deductions", "sd_check", "sd_refund_date", "sd_refund_ack_no", "date_of_refund_application_ri", "refund_application_no_ri", "ri_refund_amount", "penalty_deductions", "remarks_for_deductions_ri", "ri_refund_date", "ri_refund_ack_no"
    ]
    missing = [col for col in expected_columns if col not in df.columns]
    extra = [col for col in df.columns if col not in expected_columns]
    print("[UPLOAD MASTER TRACKER] Missing columns:", missing)
    print("[UPLOAD MASTER TRACKER] Extra columns:", extra)

    # Clean DataFrame: replace inf/-inf with NA, then fill all NA/NaN with None
    df = df.replace([np.inf, -np.inf], np.nan)
    print("[UPLOAD MASTER TRACKER] After replacing inf/-inf with nan:\n", df.head())
    df = df.where(pd.notnull(df), None)
    print("[UPLOAD MASTER TRACKER] After replacing nan with None:\n", df.head())

    # Convert all empty strings to None for non-text fields
    TEXT_FIELDS = {
        "dn_number", "circle", "state", "city", "area", "type_of_build", "financial_year", "link_name",
        "ce_route_id", "ce_section_id", "ce_subsection_id", "municipal_authority", "ward", "road_name",
        "start_point", "end_point", "authority_application_number", "authority_new_application_number",
        "differential_dn_number", "permit_no", "permitted_no_of_ducts", "work_status", "refund_application_no",
        "remarks_for_deductions", "sd_check", "sd_refund_ack_no", "refund_application_no_ri",
        "remarks_for_deductions_ri", "ri_refund_ack_no"
    }
    cleaned_rows = []
    for idx, row in df.iterrows():
        data = row.to_dict()
        for k in list(data.keys()):
            if data[k] == "" and k not in TEXT_FIELDS:
                data[k] = None
        # Clean all NaN/numpy.nan/inf recursively
        data = clean_json(data)
        cleaned_rows.append(data)
        if idx < 3:
            print(f"[UPLOAD MASTER TRACKER] Cleaned row {idx}: {data}")
    print(f"[UPLOAD MASTER TRACKER] Cleaned {len(cleaned_rows)} rows. Sample:", cleaned_rows[:2])

    # Bulk upsert all rows (using dn_number as unique key)
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    errors = []
    try:
        response = supabase.table("master_tracker").upsert(cleaned_rows, on_conflict="dn_number").execute()
        print("[UPLOAD MASTER TRACKER] Upsert response:", response)
        if hasattr(response, 'error') and response.error:
            errors.append(str(response.error))
    except Exception as e:
        print("[UPLOAD MASTER TRACKER] Exception during upsert:", e)
        errors.append(str(e))

    result = {
        "success": len(errors) == 0,
        "errors": errors,
        "rows": len(cleaned_rows),
        "cleaned_rows": cleaned_rows[:5],  # Already cleaned!
        "message": "All rows upserted successfully." if len(errors) == 0 else "Some rows failed."
    }
    print("[UPLOAD MASTER TRACKER] Final response to client:", result)
    return clean_json(result)

# Add this after FIELD_MAP
PRETTY_TO_CANONICAL_FIELD_MAP = {v: k for k, v in FIELD_MAP.items()}

@app.get("/api/route-ids")
def get_route_ids():
    """
    Returns all unique route_id_site_id values from dn_master where route_type is 'DC Route' or 'Additional Route'.
    """
    try:
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        # Query for both DC Route and Additional Route types
        response = supabase.table("dn_master").select("route_id_site_id, route_type").in_("route_type", ["DC Route", "Additional Route"]).execute()
        data = response.data or []
        route_ids = list({row["route_id_site_id"] for row in data if row.get("route_id_site_id")})
        print(f"[ROUTE_IDS] Found {len(route_ids)} unique route_id_site_id values for DC Route and Additional Route types")
        return {"route_ids": route_ids}
    except Exception as e:
        print(f"[ERROR] Failed to fetch route ids: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch route ids.")

@app.get("/api/route-site-ids")
def get_route_site_ids(route_id: str):
    """
    Returns all route_id_site_id values from budget_master where uid matches the given route_id.
    """
    try:
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        response = supabase.table("budget_master").select("route_id_site_id, uid").eq("uid", route_id).execute()
        data = response.data or []
        site_ids = [row["route_id_site_id"] for row in data if row.get("route_id_site_id")]
        return {"site_ids": site_ids}
    except Exception as e:
        print(f"[ERROR] Failed to fetch site ids for route: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch site ids for route.")

@app.get("/api/route-analysis")
def route_analysis(route_id_site_id: str):
    """
    Returns all rows in budget_master for the given route_id_site_id,
    with fields: ce_length_mtr, total_ri_amount, material_cost, execution_cost_including_hh.
    """
    try:
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        response = supabase.table("budget_master").select(
            "ce_length_mtr, total_ri_amount, material_cost, execution_cost_including_hh"
        ).eq("route_id_site_id", route_id_site_id).execute()
        data = response.data or []
        return {"data": data}
    except Exception as e:
        print(f"[ERROR] Failed to fetch route analysis: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch route analysis.")

@app.post("/api/generate-route-report")
async def generate_route_report(request: Request):
    data = await request.json()
    route_id = data.get("route_id", "")
    budget_summary = data.get("budget_summary", {})
    pre_table = data.get("pre_table", {"headers": [], "rows": []})
    current_table = data.get("current_table", {"headers": [], "rows": []})
    post_table = data.get("post_table", {"headers": [], "rows": []})
    print("[ROUTE REPORT] pre_table headers:", pre_table.get("headers"))
    print("[ROUTE REPORT] pre_table rows (first 2):", pre_table.get("rows", [])[:2])
    print("[ROUTE REPORT] current_table headers:", current_table.get("headers"))
    print("[ROUTE REPORT] current_table rows (first 2):", current_table.get("rows", [])[:2])
    print("[ROUTE REPORT] post_table headers:", post_table.get("headers"))
    print("[ROUTE REPORT] post_table rows (first 2):", post_table.get("rows", [])[:2])
    remaining = data.get("remaining_analysis", {})
    chart_image = data.get("chart_image", None)

    buffer = io.BytesIO()
    # We'll build elements for portrait and landscape separately
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph(f"<b>Route Report: {route_id}</b>", styles['Title']))
    elements.append(Spacer(1, 12))
    # Budget summary as key-value table (portrait)
    if budget_summary and isinstance(budget_summary, dict):
        summary_data = [[Paragraph(f"<b>{str(k)}</b>", styles['Normal']), str(v)] for k, v in budget_summary.items()]
        summary_table = Table(summary_data, colWidths=[120, 120])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#232f47')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ]))
        elements.append(Paragraph("<b>Budget Summary</b>", styles['Heading2']))
        elements.append(summary_table)
        elements.append(Spacer(1, 12))
    # Helper to add a table, using landscape for wide tables
    def add_table(title, table_data):
        headers = table_data["headers"]
        rows = table_data["rows"]
        if not headers or not rows:
            return
        max_cols = 10
        if len(headers) > max_cols:
            # Wide table: landscape page
            elements.append(PageBreak())
            elements.append(Paragraph(f"<b>{title}</b>", styles['Heading2']))
            data = [headers] + rows
            t = Table(data, repeatRows=1, colWidths=[max(80, int(950/len(headers)))]*len(headers))
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#232f47')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            for i in range(1, len(data)):
                if i % 2 == 0:
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#e6eaf2')),
                    ]))
            elements.append(t)
            elements.append(Spacer(1, 12))
        else:
            # Normal table: portrait
            elements.append(Paragraph(f"<b>{title}</b>", styles['Heading2']))
            data = [headers] + rows
            t = Table(data, repeatRows=1, colWidths=[max(60, int(700/len(headers)))]*len(headers))
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#232f47')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            for i in range(1, len(data)):
                if i % 2 == 0:
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#e6eaf2')),
                    ]))
            elements.append(t)
            elements.append(Spacer(1, 12))
    # Add tables
    add_table("Pre Analysis Table", pre_table)
    add_table("Current Analysis Table", current_table)
    add_table("Post Analysis Table", post_table)
    # Remaining Route Analysis
    elements.append(Paragraph("<b>Remaining Route Analysis</b>", styles['Heading2']))
    if isinstance(remaining, dict):
        rem_data = [[Paragraph(f"<b>{str(k)}</b>", styles['Normal']), str(v)] for k, v in remaining.items()]
        rem_table = Table(rem_data, colWidths=[120, 120])
        rem_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#232f47')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ]))
        elements.append(rem_table)
    else:
        elements.append(Paragraph(str(remaining), styles['Normal']))
    elements.append(Spacer(1, 12))
    # S Curve Chart
    if chart_image:
        elements.append(PageBreak())
        try:
            elements.append(Paragraph("<b>S Curve Chart (Budget vs Actuals)</b>", styles['Heading2']))
            imgdata = chart_image.split(",", 1)[-1]
            img_bytes = base64.b64decode(imgdata)
            from reportlab.platypus import Image
            img = Image(io.BytesIO(img_bytes), width=400, height=180)
            elements.append(img)
            elements.append(Spacer(1, 18))
        except Exception as e:
            elements.append(Paragraph(f"[Error rendering chart image: {e}]", styles['Normal']))
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=18, leftMargin=18, topMargin=24, bottomMargin=18)
    doc.build(elements)
    buffer.seek(0)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(buffer.read())
        tmp_path = tmp.name
    return FileResponse(tmp_path, filename=f"Route_Report_{route_id}.pdf", media_type="application/pdf")

# --- KDMC Extraction Endpoint ---
@app.post("/api/kdmc-extract")
async def kdmc_extract(file: UploadFile = File(...)):
    import tempfile, os
    from parsers import kdmc
    from constants.comprehensive_field_mapping import convert_standard_to_table, ensure_all_fields_present
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = os.path.join(temp_dir, file.filename)
        with open(temp_path, "wb") as tmp:
            tmp.write(await file.read())
        # Use the new comprehensive parser with hardcoded values
        non_refundable_result = kdmc.non_refundable_request_parser(temp_path)
        # Get SD data using the extracted fields
        all_fields = kdmc.extract_kdmc_all_fields(temp_path)
        sd_headers, sd_row = kdmc.sd_parser(temp_path)
        
        return {
            "non_refundable": non_refundable_result,
            "sd_headers": sd_headers,
            "sd_row": sd_row,
            "all_fields": all_fields
        }

# --- Client Parser V2 Endpoints ---
@app.post("/api/client-parser-v2/unified")
async def client_parser_v2_unified(
    dn_number: str = Form(...),
    authority: str = Form(...),
    output_type: str = Form("both")  # "non_refundable", "sd", or "both"
):
    """
    Unified endpoint for Client Parser V2 - generates both non-refundable and SD outputs
    based on DN number and authority without any PDF parsing.
    """
    try:
        logger.info(f"Client Parser V2 - Unified request for DN: {dn_number}, Authority: {authority}, Type: {output_type}")
        
        # Validate inputs
        if not dn_number or not authority:
            raise HTTPException(status_code=400, detail="DN number and authority are required")
        
        # Call the unified parser
        result = unified_parser(dn_number, authority, output_type)
        
        if "error" in result:
            raise HTTPException(status_code=500, detail=result["error"])
        
        logger.info(f"Client Parser V2 - Unified completed successfully")
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in client parser v2 unified: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.post("/api/client-parser-v2/non-refundable")
async def client_parser_v2_non_refundable(
    dn_number: str = Form(...),
    authority: str = Form(...)
):
    """
    Client Parser V2 - Non-refundable output only
    """
    try:
        logger.info(f"Client Parser V2 - Non-refundable request for DN: {dn_number}, Authority: {authority}")
        
        # Validate inputs
        if not dn_number or not authority:
            raise HTTPException(status_code=400, detail="DN number and authority are required")
        
        # Call the non-refundable parser
        result = generate_non_refundable_output(dn_number, authority)
        
        logger.info(f"Client Parser V2 - Non-refundable completed successfully")
        return {"non_refundable": result}
        
    except Exception as e:
        logger.error(f"Error in client parser v2 non-refundable: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.post("/api/client-parser-v2/sd")
async def client_parser_v2_sd(
    dn_number: str = Form(...),
    authority: str = Form(...)
):
    """
    Client Parser V2 - SD output only
    """
    try:
        logger.info(f"Client Parser V2 - SD request for DN: {dn_number}, Authority: {authority}")
        
        # Validate inputs
        if not dn_number or not authority:
            raise HTTPException(status_code=400, detail="DN number and authority are required")
        
        # Call the SD parser
        result = generate_sd_output(dn_number, authority)
        
        logger.info(f"Client Parser V2 - SD completed successfully")
        return {"sd": result}
        
    except Exception as e:
        logger.error(f"Error in client parser v2 sd: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.get("/api/client-parser-v2/authorities")
async def get_available_authorities():
    """
    Get list of available authorities for Client Parser V2
    """
    try:
        from parsers.clientparserv2 import AUTHORITY_CONFIGS
        authorities = list(AUTHORITY_CONFIGS.keys())
        return {"authorities": authorities}
    except Exception as e:
        logger.error(f"Error getting authorities: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.get("/api/client-parser-v2/dn-numbers")
async def get_dn_numbers():
    """
    Get all DN numbers from the database for the dropdown
    """
    try:
        logger.info("Fetching all DN numbers from database")
        
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        
        # Query all DN numbers from the dn_master table
        response = supabase.table("dn_master").select("dn_number").execute()
        
        if response.data:
            dn_numbers = [row["dn_number"] for row in response.data if row.get("dn_number")]
            logger.info(f"Found {len(dn_numbers)} DN numbers")
            return {"dn_numbers": dn_numbers}
        else:
            logger.warning("No DN numbers found in database")
            return {"dn_numbers": []}
            
    except Exception as e:
        logger.error(f"Error fetching DN numbers: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.post("/api/client-parser-v2/validate-dn")
async def validate_dn_number(
    dn_number: str = Form(...)
):
    """
    Validate if a DN number exists in the database
    """
    try:
        logger.info(f"Validating DN number: {dn_number}")
        
        from parsers.clientparserv2 import query_dn_master
        
        # Query the database
        dn_data = query_dn_master(dn_number)
        
        if dn_data:
            return {
                "exists": True,
                "dn_number": dn_number,
                "message": "DN number found in database"
            }
        else:
            return {
                "exists": False,
                "dn_number": dn_number,
                "message": "DN number not found in database"
            }
            
    except Exception as e:
        logger.error(f"Error validating DN number: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
